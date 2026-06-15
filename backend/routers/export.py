import contextlib
import datetime
import io
from io import BytesIO
from urllib.parse import quote

import qrcode
from fastapi import APIRouter, Depends, HTTPException, Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    Image,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy import and_, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from backend import crud
from backend.auth import jwt_utils
from backend.auth.module_guard import require_module_dep
from backend.database.database import get_db
from backend.database.models import (
    Bonus,
    Company,
    ContractAnnex,
    EmploymentContract,
    Invoice,
    InvoiceItem,
    LeaveRequest,
    Payslip,
    TimeLog,
    User,
    WorkSchedule,
)
from backend.services.payroll_calculator import PayrollCalculator

router = APIRouter(
    prefix="/export",
    tags=["export"],
    dependencies=[Depends(require_module_dep("salaries"))],
)

# Register Fonts for Cyrillic support
try:
    pdfmetrics.registerFont(TTFont("DejaVuSans", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"))
    pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"))
    DEFAULT_FONT = "DejaVuSans"
    BOLD_FONT = "DejaVuSans-Bold"
except Exception as e:
    print(f"Warning: Could not register Cyrillic fonts: {e}")
    DEFAULT_FONT = "Helvetica"
    BOLD_FONT = "Helvetica-Bold"

@router.get("/payroll/xlsx")
async def export_payroll_xlsx(
    start_date: datetime.date,
    end_date: datetime.date,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(jwt_utils.get_current_user),
):
    """Exports a comprehensive payroll report for ALL active users in the given period to Excel.
    """
    if current_user.role.name not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Only admins can export full payroll data")

    # 1. Fetch all active users
    result = await db.execute(select(User).where(User.is_active))
    users = result.scalars().all()

    # 2. Setup Excel Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Ведомост"

    # Header Row
    headers = [
        "Служител", "Имейл", "Длъжност", "Редовни часове", "Извънредни часове",
        "Болнични (дни)", "Отпуск (дни)", "Сума Редовни", "Сума Извънредни",
        "Бонуси", "Бруто", "Осигуровки", "ДДФЛ", "НЕТО ЗА ПОЛУЧАВАНЕ", "Валута",
    ]
    ws.append(headers)

    # Style Header
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 3. Calculate data for each user
    calculator = PayrollCalculator(db)
    # Convert dates to datetime for calculator
    start_dt = datetime.datetime.combine(start_date, datetime.time.min)
    end_dt = datetime.datetime.combine(end_date, datetime.time.max)

    for employee in users:
        try:
            res = await calculator.calculate(employee.id, start_dt, end_dt)

            # Get currency
            payroll_config = await crud.get_payroll_config(db, employee.id)
            currency = payroll_config.currency if hasattr(payroll_config, "currency") else "EUR"

            # Aggregated totals
            gross = res["regular_amount"] + res["overtime_amount"] + res["bonus_amount"]

            ws.append([
                f"{employee.first_name} {employee.last_name}",
                employee.email,
                employee.job_title or "-",
                res["total_regular_hours"],
                res["total_overtime_hours"],
                res["sick_days"],
                res["leave_days"],
                res["regular_amount"],
                res["overtime_amount"],
                res["bonus_amount"],
                round(gross, 2),
                res["insurance_amount"],
                res["tax_amount"],
                res["total_amount"],
                currency,
            ])
        except Exception as e:
            # Skip or log error for specific user to not break full export
            ws.append([f"{employee.first_name} {employee.last_name}", employee.email, "ГРЕШКА ПРИ ИЗЧИСЛЕНИЕ", str(e)])

    # Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            with contextlib.suppress(BaseException):
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    # 4. Stream response
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_filename = f"payroll_{start_date}_{end_date}.xlsx"
    encoded_filename = quote(safe_filename)

    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=\"{safe_filename}\"; filename*=UTF-8''{encoded_filename}",
        },
    )


@router.get("/annex/{annex_id}/pdf")
async def export_annex_pdf(
    annex_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(jwt_utils.get_current_user),
):
    """Генерира PDF на допълнително споразумение."""
    if current_user.role.name not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Only admins can export annexes")

    result = await db.execute(select(ContractAnnex).where(ContractAnnex.id == annex_id))
    annex = result.scalars().first()
    if not annex:
        raise HTTPException(status_code=404, detail="Annex not found")

    contract_result = await db.execute(
        select(EmploymentContract).where(EmploymentContract.id == annex.contract_id),
    )
    contract = contract_result.scalars().first()
    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=50, rightMargin=50, topMargin=50, bottomMargin=50)
    elements = []
    styles = getSampleStyleSheet()

    for style_name in styles.byName:
        style = styles[style_name]
        if "Bold" in style_name or "Title" in style_name or "Heading" in style_name:
            style.fontName = BOLD_FONT
        else:
            style.fontName = DEFAULT_FONT

    if "Italic" in styles:
        styles["Italic"].fontName = DEFAULT_FONT
    else:
        styles.add(styles["Normal"].clone("Italic"))
        styles["Italic"].fontName = DEFAULT_FONT

    elements.append(Paragraph("ДОПЪЛНИТЕЛНО СПОРАЗУМЕНИЕ", styles["Title"]))
    elements.append(Spacer(1, 8))
    elements.append(Paragraph("Към трудов договор", styles["Normal"]))
    elements.append(Spacer(1, 16))

    if contract.company:
        elements.append(Paragraph(f"<b>{contract.company.name}</b>", styles["Normal"]))
        if contract.company.address:
            elements.append(Paragraph(f"Адрес: {contract.company.address}", styles["Normal"]))
        if contract.company.mol_name:
            elements.append(Paragraph(f"Представляван от: {contract.company.mol_name}", styles["Normal"]))
        elements.append(Spacer(1, 12))

    if contract.contract_number:
        elements.append(Paragraph(f"<b>Към Договор №:</b> {contract.contract_number}", styles["Normal"]))
    if annex.annex_number:
        elements.append(Paragraph(f"<b>Споразумение №:</b> {annex.annex_number}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Дата на влизане в сила:</b> {annex.effective_date.strftime('%d.%m.%Y') if annex.effective_date else 'N/A'}", styles["Normal"]))

    elements.append(Spacer(1, 20))
    elements.append(Paragraph("<b>СТРАНИ ПО СПОРАЗУМЕНИЕТО</b>", styles["Heading2"]))
    elements.append(Spacer(1, 8))

    employee_name = contract.employee_name or (current_user.first_name + " " + current_user.last_name if contract.user_id == current_user.id else "N/A")
    employee_egn = contract.employee_egn or (current_user.egn if contract.user_id == current_user.id else "N/A")

    elements.append(Paragraph(f"<b>Работодател:</b> {contract.company.name if contract.company else 'N/A'}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Работник:</b> {employee_name}, ЕГН: {employee_egn}", styles["Normal"]))

    elements.append(Spacer(1, 20))
    elements.append(Paragraph("<b>ИЗМЕНЕНИЯ И ДОПЪЛНЕНИЯ</b>", styles["Heading2"]))
    elements.append(Spacer(1, 8))

    if annex.change_type:
        change_type_display = {
            "salary_increase": "Повишение на основната месечна заплата",
            "salary_decrease": "Намаление на основната месечна заплата",
            "position_change": "Промяна на длъжността",
            "hours_change": "Промяна на работното време",
            "other": "Други изменения",
        }.get(annex.change_type, annex.change_type)
        elements.append(Paragraph(f"<b>Вид изменение:</b> {change_type_display}", styles["Normal"]))

    if annex.change_description:
        elements.append(Spacer(1, 8))
        elements.append(Paragraph(f"<b>Описание:</b> {annex.change_description}", styles["Normal"]))

    if annex.base_salary:
        elements.append(Spacer(1, 8))
        elements.append(Paragraph(f"<b>Нова основна месечна заплата:</b> {float(annex.base_salary):.2f} €", styles["Normal"]))

    if annex.work_hours_per_week:
        elements.append(Spacer(1, 8))
        elements.append(Paragraph(f"<b>Ново работно време:</b> {annex.work_hours_per_week} часа седмично", styles["Normal"]))

    if annex.night_work_rate:
        elements.append(Spacer(1, 8))
        elements.append(Paragraph(f"<b>Множител за нощен труд:</b> {annex.night_work_rate}", styles["Normal"]))

    if annex.overtime_rate:
        elements.append(Paragraph(f"<b>Множител за извънреден труд:</b> {annex.overtime_rate}", styles["Normal"]))

    if annex.holiday_rate:
        elements.append(Paragraph(f"<b>Множител за работа на официални празници:</b> {annex.holiday_rate}", styles["Normal"]))

    elements.append(Spacer(1, 24))
    elements.append(Paragraph("<b>ПОДПИСИ</b>", styles["Heading2"]))
    elements.append(Spacer(1, 36))

    elements.append(Paragraph("_" * 40 + "            " + "_" * 40, styles["Normal"]))
    elements.append(Paragraph("Работодател                                      Работник", styles["Normal"]))
    elements.append(Spacer(1, 12))

    if annex.signed_at:
        elements.append(Paragraph(f"<i>Подписано на: {annex.signed_at.strftime('%d.%m.%Y %H:%M')}</i>", styles["Normal"]))
    elif annex.status == "signed":
        elements.append(Paragraph("<i>(Очаква подписване)</i>", styles["Normal"]))

    doc.build(elements)

    buffer.seek(0)
    safe_filename = f"annex_{annex_id}.pdf"
    encoded_filename = quote(safe_filename)

    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=\"{safe_filename}\"; filename*=UTF-8''{encoded_filename}",
        },
    )

@router.get("/payslip/{payslip_id}/pdf")
async def export_payslip_pdf(
    payslip_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(jwt_utils.get_current_user),
):
    # 1. Fetch Payslip
    result = await db.execute(select(Payslip).where(Payslip.id == payslip_id))
    payslip = result.scalars().first()
    if not payslip:
        raise HTTPException(status_code=404, detail="Payslip not found")

    # 2. Permissions Check
    if current_user.role.name not in ["admin", "super_admin"] and current_user.id != payslip.user_id:
        raise HTTPException(status_code=403, detail="Access denied")

    # 3. Fetch User & Data
    user_res = await db.execute(select(User).where(User.id == payslip.user_id))
    employee = user_res.scalars().first()

    # Fetch Logs
    logs_res = await db.execute(
        select(TimeLog)
        .where(TimeLog.user_id == payslip.user_id)
        .where(TimeLog.start_time >= payslip.period_start)
        .where(TimeLog.start_time <= payslip.period_end)
        .order_by(TimeLog.start_time.asc()),
    )
    logs = logs_res.scalars().all()

    # Fetch Bonuses
    bonus_res = await db.execute(
        select(Bonus)
        .where(Bonus.user_id == payslip.user_id)
        .where(Bonus.date >= payslip.period_start.date())
        .where(Bonus.date <= payslip.period_end.date()),
    )
    bonuses = bonus_res.scalars().all()

    # Fetch Leaves
    leave_res = await db.execute(
        select(LeaveRequest)
        .where(LeaveRequest.user_id == payslip.user_id)
        .where(LeaveRequest.status == "approved")
        .where(or_(
            and_(LeaveRequest.start_date >= payslip.period_start.date(), LeaveRequest.start_date <= payslip.period_end.date()),
            and_(LeaveRequest.end_date >= payslip.period_start.date(), LeaveRequest.end_date <= payslip.period_end.date()),
        )),
    )
    leaves = leave_res.scalars().all()

    # Fetch Schedules
    sched_res = await db.execute(
        select(WorkSchedule)
        .where(WorkSchedule.user_id == payslip.user_id)
        .where(WorkSchedule.date >= payslip.period_start.date())
        .where(WorkSchedule.date <= payslip.period_end.date())
        .options(selectinload(WorkSchedule.shift)),
    )
    schedules_data = {s.date: s.shift for s in sched_res.scalars().all() if s.shift}

    # 4. Fetch Currency
    payroll_config = await crud.get_payroll_config(db, payslip.user_id)
    currency = payroll_config.currency if hasattr(payroll_config, "currency") else "EUR"

    # 5. Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    # Update ALL styles to use Cyrillic-compatible font
    for style_name in styles.byName:
        style = styles[style_name]
        if "Bold" in style_name or "Title" in style_name or "Heading" in style_name:
            style.fontName = BOLD_FONT
        else:
            style.fontName = DEFAULT_FONT

    # Standard stylesheet already has 'Italic', just update its font
    if "Italic" in styles:
        styles["Italic"].fontName = DEFAULT_FONT
    else:
        # Fallback in case it's missing in some environments
        styles.add(styles["Normal"].clone("Italic"))
        styles["Italic"].fontName = DEFAULT_FONT

    # Title
    elements.append(Paragraph("ФИШ ЗА ЗАПЛАТА / ДЕТАЙЛЕН ОТЧЕТ", styles["Title"]))
    elements.append(Spacer(1, 12))

    # Info
    elements.append(Paragraph(f"<b>Служител:</b> {employee.first_name} {employee.last_name} ({employee.email})", styles["Normal"]))
    elements.append(Paragraph(f"<b>Период:</b> {payslip.period_start.date()} до {payslip.period_end.date()}", styles["Normal"]))
    elements.append(Spacer(1, 20))

    # --- TABLE DATA PREPARATION ---
    # Merge logs and leaves into a chronological list of events
    events = []

    for log in logs:
        events.append({
            "date": log.start_time.date(),
            "type": "log",
            "obj": log,
        })

    # Expand leaves into days
    from datetime import timedelta
    for leave in leaves:
        curr = leave.start_date
        while curr <= leave.end_date:
            if curr >= payslip.period_start.date() and curr <= payslip.period_end.date() and curr.weekday() < 5:
                events.append({
                    "date": curr,
                    "type": "leave",
                    "obj": leave,
                })
            curr += timedelta(days=1)

    # Sort events
    events.sort(key=lambda x: (x["date"], x["obj"].start_time if x["type"] == "log" else datetime.time.min))

    # Table Header
    data = [["Дата", "Смяна", "Тип", "Вход", "Изход", "Почивка", "Извънр.", "Общо"]]

    for event in events:
        date = event["date"]
        shift_obj = schedules_data.get(date)
        shift_name = shift_obj.name if shift_obj else "-"

        if event["type"] == "log":
            log = event["obj"]
            entry_type = "Ръчен" if log.is_manual else "Авт."

            in_time = log.start_time.strftime("%H:%M")
            out_time = log.end_time.strftime("%H:%M") if log.end_time else "Активен"

            duration = datetime.timedelta()
            ot_seconds = 0

            if log.end_time:
                duration = log.end_time - log.start_time
                if log.break_duration_minutes:
                    duration -= datetime.timedelta(minutes=log.break_duration_minutes)

                net_duration_seconds = duration.total_seconds()

                if shift_obj and shift_obj.shift_type == "regular":
                    s_start = datetime.datetime.combine(date, shift_obj.start_time)
                    s_end = datetime.datetime.combine(date, shift_obj.end_time)
                    if s_end < s_start: s_end += timedelta(days=1)

                    overlap_start = max(log.start_time, s_start)
                    overlap_end = min(log.end_time, s_end)

                    overlap_seconds = 0
                    if overlap_end > overlap_start:
                        overlap_seconds = (overlap_end - overlap_start).total_seconds()

                    reg_seconds = max(0, overlap_seconds - (log.break_duration_minutes * 60))
                    ot_seconds = max(0, net_duration_seconds - reg_seconds)
                else:
                    ot_seconds = net_duration_seconds

            hours, remainder = divmod(duration.total_seconds(), 3600)
            minutes, _ = divmod(remainder, 60)
            dur_str = f"{int(hours)}ч {int(minutes)}м"

            ot_h, ot_rem = divmod(ot_seconds, 3600)
            ot_m, _ = divmod(ot_rem, 60)
            ot_str = f"{int(ot_h)}ч {int(ot_m)}м" if ot_seconds > 0 else "-"

            data.append([
                date.strftime("%Y-%m-%d"),
                shift_name,
                entry_type,
                in_time,
                out_time,
                f"{log.break_duration_minutes}м",
                ot_str,
                dur_str,
            ])

        elif event["type"] == "leave":
            leave = event["obj"]
            leave_name = "Платен" if leave.leave_type == "paid_leave" else ("Болничен" if leave.leave_type == "sick_leave" else "Неплатен")
            data.append([
                date.strftime("%Y-%m-%d"),
                shift_name,
                leave_name,
                "-", "-", "-", "-", "8ч 00м",
            ])

    # Style Table
    t = Table(data, colWidths=[65, 75, 55, 45, 45, 50, 55, 55])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), BOLD_FONT),
        ("FONTNAME", (0, 0), (-1, -1), DEFAULT_FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 20))

    # --- SUMMARY SECTION ---
    elements.append(Paragraph("<b>ОБОБЩЕНИЕ</b>", styles["Heading2"]))

    # Calc Leave Days
    paid_days = sum(1 for e in events if e["type"] == "leave" and e["obj"].leave_type == "paid_leave")
    sick_days = sum(1 for e in events if e["type"] == "leave" and e["obj"].leave_type == "sick_leave")

    elements.append(Paragraph(f"<b>Редовни часове:</b> {payslip.total_regular_hours} ч.", styles["Normal"]))
    elements.append(Paragraph(f"<b>Извънредни часове:</b> {payslip.total_overtime_hours} ч.", styles["Normal"]))
    elements.append(Paragraph(f"<b>Използван платен отпуск:</b> {paid_days} дни", styles["Normal"]))
    elements.append(Paragraph(f"<b>Болнични дни:</b> {sick_days} дни", styles["Normal"]))

    total_bonus = sum(b.amount for b in bonuses)
    if total_bonus > 0:
        elements.append(Paragraph(f"<b>Бонуси:</b> {total_bonus:.2f} {currency}", styles["Normal"]))
        for b in bonuses:
             elements.append(Paragraph(f"&nbsp;&nbsp;&nbsp;- {b.date}: {b.amount:.2f} ({b.description or 'N/A'})", styles["Italic"]))

    elements.append(Spacer(1, 10))
    elements.append(Paragraph(f"<b>СУМА ЗА ПОЛУЧАВАНЕ:</b> {payslip.total_amount} {currency}", styles["Heading2"]))

    elements.append(Spacer(1, 40))
    elements.append(Paragraph(f"Генерирано на: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Italic"]))

    doc.build(elements)

    buffer.seek(0)
    import re
    ascii_last_name = re.sub(r"[^\x00-\x7F]+", "", employee.last_name) or "employee"
    safe_filename = f"report_{ascii_last_name}_{payslip_id}.pdf"
    full_filename = f"report_{employee.last_name}_{payslip_id}.pdf"
    encoded_filename = quote(full_filename)

    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=\"{safe_filename}\"; filename*=UTF-8''{encoded_filename}",
        },
    )


@router.get("/contract/{contract_id}/pdf")
async def export_contract_pdf(
    contract_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(jwt_utils.get_current_user),
):
    if current_user.role.name not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Only admins can export contracts")

    result = await db.execute(
        select(EmploymentContract).options(
            selectinload(EmploymentContract.company),
            selectinload(EmploymentContract.position),
            selectinload(EmploymentContract.department),
            selectinload(EmploymentContract.user),
        ).where(EmploymentContract.id == contract_id)
    )
    contract = result.scalars().first()
    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")

    sections = []
    clauses = []
    if contract.template_id:
        from backend.database.models import (
            ContractTemplate,
            ContractTemplateClause,
            ContractTemplateSection,
            ContractTemplateVersion,
            ClauseTemplate,
        )
        template = await db.get(ContractTemplate, contract.template_id)
        if template:
            ver_result = await db.execute(
                select(ContractTemplateVersion).where(
                    ContractTemplateVersion.template_id == contract.template_id,
                    ContractTemplateVersion.is_current,
                )
            )
            current_version = ver_result.scalar_one_or_none()
            if current_version:
                sec_result = await db.execute(
                    select(ContractTemplateSection)
                    .where(ContractTemplateSection.version_id == current_version.id)
                    .order_by(ContractTemplateSection.order_index)
                )
                sections = sec_result.scalars().all()

        clause_result = await db.execute(
            select(ContractTemplateClause)
            .where(ContractTemplateClause.template_id == contract.template_id)
            .order_by(ContractTemplateClause.order_index)
        )
        template_clause_links = clause_result.scalars().all()
        for link in template_clause_links:
            clause = await db.get(ClauseTemplate, link.clause_id)
            if clause:
                clauses.append(clause)

    if contract.clause_ids:
        import json
        try:
            extra_ids = json.loads(contract.clause_ids)
            existing_ids = {c.id for c in clauses}
            from backend.database.models import ClauseTemplate
            for cid in extra_ids:
                if cid not in existing_ids:
                    c = await db.get(ClauseTemplate, cid)
                    if c:
                        clauses.append(c)
        except (json.JSONDecodeError, TypeError):
            pass

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=50, rightMargin=50, topMargin=50, bottomMargin=50)
    elements = []
    styles = getSampleStyleSheet()

    for style_name in styles.byName:
        style = styles[style_name]
        if "Bold" in style_name or "Title" in style_name or "Heading" in style_name:
            style.fontName = BOLD_FONT
        else:
            style.fontName = DEFAULT_FONT

    if "Italic" in styles:
        styles["Italic"].fontName = DEFAULT_FONT
    else:
        styles.add(styles["Normal"].clone("Italic"))
        styles["Italic"].fontName = DEFAULT_FONT

    Normal = styles["Normal"]
    Title = styles["Title"]
    Heading2 = styles["Heading2"]
    Italic = styles["Italic"]

    contract_type_display = {
        "full_time": "ТРУДОВ ДОГОВОР ЗА НЕОПРЕДЕЛЕНО РАБОТНО ВРЕМЕ",
        "part_time": "ТРУДОВ ДОГОВОР ЗА ОПРЕДЕЛЕНО РАБОТНО ВРЕМЕ",
        "contractor": "ГРАЖДАНСКИ ДОГОВОР",
        "internship": "ДОГОВОР ЗА СТАЖ",
    }.get(contract.contract_type, "ТРУДОВ ДОГОВОР")

    contract_num = contract.contract_number or "______"
    sign_date = contract.start_date.strftime("%d.%m.%Y") if contract.start_date else "................"

    elements.append(Paragraph(f"<b>Т Р У Д О В Д О Г О В О Р № {contract_num}</b>", Title))
    elements.append(Spacer(1, 18))

    elements.append(Paragraph(
        f"Днес, {sign_date} между следните страни:",
        Normal,
    ))
    elements.append(Spacer(1, 12))

    company = contract.company
    elements.append(Paragraph("<b>РАБОТОДАТЕЛ:</b>", Heading2))
    if company:
        elements.append(Paragraph(f"<b>{company.name}</b>", Normal))
        if company.address:
            elements.append(Paragraph(f"Адрес: {company.address}", Normal))
        if company.eik:
            elements.append(Paragraph(f"ЕИК по Булстат: {company.eik}", Normal))
        if company.mol_name:
            elements.append(Paragraph(f"Представлявано от: {company.mol_name}, Управител", Normal))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("от една страна, наричана за краткост по-долу <b>РАБОТОДАТЕЛ</b>, и", Normal))
    elements.append(Spacer(1, 12))

    employee = contract.user
    emp_name = contract.employee_name or (f"{employee.first_name or ''} {employee.surname or ''} {employee.last_name or ''}".strip() if employee else "................................")
    emp_egn = contract.employee_egn or (employee.egn if employee else "............")
    emp_address = (employee.address if employee and employee.address else "................................")

    elements.append(Paragraph("<b>РАБОТНИК:</b>", Heading2))
    elements.append(Paragraph(f"{emp_name}, ЕГН: {emp_egn}", Normal))
    elements.append(Paragraph(f"Адрес: {emp_address}", Normal))
    if employee:
        if employee.phone_number:
            elements.append(Paragraph(f"Телефон: {employee.phone_number}", Normal))
    elements.append(Spacer(1, 6))

    elements.append(Paragraph(
        "наричан за краткост по-долу <b>РАБОТНИК</b>, на основание чл. 67, ал. 1, т. 1 от Кодекса на труда "
        "се сключи настоящият трудов договор.",
        Normal,
    ))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph(
        "Работодателят възлага, а работникът приема да изпълнява трудовите си задължения при следните условия:",
        Normal,
    ))
    elements.append(Spacer(1, 12))

    section_num = 1

    position = contract.position
    dept = contract.department
    pos_title = position.title if position else "................................"
    dept_name = dept.name if dept else "................................"

    elements.append(Paragraph(
        f"<b>{section_num}. Длъжност:</b> „{pos_title}",
        Normal,
    ))
    elements.append(Paragraph(f"Отдел: {dept_name}", Normal))
    elements.append(Paragraph(
        "По длъжностна характеристика, с която работникът е запознат и му е връчен екземпляр от нея.",
        Normal,
    ))
    elements.append(Spacer(1, 8))
    section_num += 1

    salary = contract.base_salary
    salary_str = f"{salary} лв." if salary else "................ лв."
    elements.append(Paragraph(
        f"<b>{section_num}. Основно месечно трудово възнаграждение</b> в размер на {salary_str}",
        Normal,
    ))
    if contract.salary_calculation_type:
        elements.append(Paragraph(
            f"Тип изчисление: {contract.salary_calculation_type}",
            Normal,
        ))
    elements.append(Spacer(1, 8))
    section_num += 1

    elements.append(Paragraph(
        f"<b>{section_num}. Допълнителни трудови възнаграждения</b> в съответствие с разпоредбите на Кодекса на труда.",
        Normal,
    ))
    if contract.night_work_rate and contract.night_work_rate > 0:
        elements.append(Paragraph(
            f"- Нощен труд: +{int(float(contract.night_work_rate) * 100)}% надбавка",
            Normal,
        ))
    if contract.overtime_rate and contract.overtime_rate > 1:
        pct = int((float(contract.overtime_rate) - 1) * 100)
        elements.append(Paragraph(
            f"- Извънреден труд: +{pct}% множител",
            Normal,
        ))
    if contract.holiday_rate and contract.holiday_rate > 1:
        pct = int((float(contract.holiday_rate) - 1) * 100)
        elements.append(Paragraph(
            f"- Работа на официални празници: +{pct}% множител",
            Normal,
        ))
    if contract.work_class:
        elements.append(Paragraph(
            f"- Трудов клас: {contract.work_class}",
            Normal,
        ))
    if contract.dangerous_work:
        elements.append(Paragraph(
            "- Вредни условия на труд: Да",
            Normal,
        ))
    elements.append(Spacer(1, 8))
    section_num += 1

    payment_day = contract.payment_day or 25
    elements.append(Paragraph(
        f"<b>{section_num}. Възнагражденията се изплащат</b> еднократно всеки месец, до {payment_day}-то число на месеца, следващ този, за който се дължат.",
        Normal,
    ))
    elements.append(Spacer(1, 8))
    section_num += 1

    hours = contract.work_hours_per_week or 40
    daily = hours / 5
    hours_label = "Пълно работно време" if hours >= 40 else "Непълно работно време"
    elements.append(Paragraph(
        f"<b>{section_num}. Продължителност, разпределение и отчитане на работното време:</b>",
        Normal,
    ))
    elements.append(Paragraph(
        f"- {hours_label}: {daily:.1f} часа дневно, {hours} часа седмично;",
        Normal,
    ))
    elements.append(Paragraph(
        "- Разпределение на работното време: от понеделник до петък;",
        Normal,
    ))
    elements.append(Paragraph(
        "- Почивки: в съответствие с разпоредбите на Кодекса на труда.",
        Normal,
    ))
    elements.append(Spacer(1, 8))
    section_num += 1

    if contract.end_date:
        duration_text = f"за определен срок до {contract.end_date.strftime('%d.%m.%Y')}"
    else:
        duration_text = "за неопределено време"
    probation_text = f", със срок на изпитване {contract.probation_months} месеца" if contract.probation_months and contract.probation_months > 0 else ""
    elements.append(Paragraph(
        f"<b>{section_num}. Трудовият договор се сключва</b> {duration_text}{probation_text}.",
        Normal,
    ))
    elements.append(Spacer(1, 8))
    section_num += 1

    if contract.probation_months and contract.probation_months > 0:
        beneficiary_label = "работника" if contract.probation_beneficiary == "employee" else "работодателя"
        elements.append(Paragraph(
            f"<b>{section_num}. Срокът на изпитване</b> по трудовия договор е уговорен в полза на {beneficiary_label}.",
            Normal,
        ))
        elements.append(Spacer(1, 4))
        section_num += 1
        elements.append(Paragraph(
            f"<b>{section_num}.</b> До изтичане на срока на изпитване страната, в чиято полза е уговорен, може да прекрати договора без да дължи предизвестие.",
            Normal,
        ))
        elements.append(Spacer(1, 4))
        section_num += 1
        notice_days = contract.notice_period_days or 30
        elements.append(Paragraph(
            f"<b>{section_num}.</b> След изтичане на срока на изпитване, трудовият договор се смята за окончателно сключен и продължава действието си, като се прекратява съгласно общите разпоредби на Кодекса на труда и с еднакъв срок на предизвестие за двете страни от {notice_days} дни.",
            Normal,
        ))
        elements.append(Spacer(1, 8))
        section_num += 1

    elements.append(Paragraph(
        f"<b>{section_num}. Основен платен годишен отпуск</b> на основание чл. 155 от Кодекса на труда — 20 работни дни.",
        Normal,
    ))
    elements.append(Spacer(1, 8))
    section_num += 1

    elements.append(Paragraph(
        f"<b>{section_num}. Работникът ще постъпи на работа на:</b> {contract.start_date.strftime('%d.%m.%Y') if contract.start_date else '................'}",
        Normal,
    ))
    elements.append(Spacer(1, 12))
    section_num += 1

    for sec in sections:
        elements.append(Paragraph(
            f"<b>{section_num}. {sec.title}</b>",
            Normal,
        ))
        if sec.content:
            elements.append(Paragraph(sec.content, Normal))
        elements.append(Spacer(1, 8))
        section_num += 1

    for clause in clauses:
        elements.append(Paragraph(
            f"<b>{section_num}. {clause.title}</b>",
            Normal,
        ))
        if clause.content:
            elements.append(Paragraph(clause.content, Normal))
        elements.append(Spacer(1, 8))
        section_num += 1

    elements.append(Spacer(1, 12))
    elements.append(Paragraph(
        "Настоящият трудов договор се сключи в два еднообразни екземпляра, по един за всяка от страните. "
        "За неуредените в настоящия трудов договор условия се прилагат общите разпоредби на Кодекса на труда "
        "и нормативните актове на българското законодателство.",
        Normal,
    ))
    elements.append(Spacer(1, 36))

    sign_table = Table(
        [
            [Paragraph("<b>РАБОТОДАТЕЛ:</b>", Normal), Paragraph("", Normal), Paragraph("<b>РАБОТНИК:</b>", Normal)],
            [Paragraph("________________________", Normal), Paragraph("", Normal), Paragraph("________________________", Normal)],
            [Paragraph(f"{company.mol_name if company and company.mol_name else ''}", Normal), Paragraph("", Normal), Paragraph(f"{emp_name}", Normal)],
        ],
        colWidths=[200, 60, 200],
    )
    sign_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
    ]))
    elements.append(sign_table)
    elements.append(Spacer(1, 24))

    elements.append(Paragraph(
        "Екземпляр от трудовия договор, заверено Уведомление по чл. 62, ал. 5 от Кодекса на труда до ТД на НАП "
        "и длъжностна характеристика са връчени на работника преди постъпване на работа.",
        Normal,
    ))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(
        f"Работникът постъпи на работа на: {contract.start_date.strftime('%d.%m.%Y') if contract.start_date else '................'}",
        Normal,
    ))

    if contract.signed_at:
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"<i>Подписан на: {contract.signed_at.strftime('%d.%m.%Y %H:%M')}</i>", Italic))

    doc.build(elements)

    buffer.seek(0)
    safe_filename = f"contract_{contract_id}.pdf"
    encoded_filename = quote(safe_filename)

    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=\"{safe_filename}\"; filename*=UTF-8''{encoded_filename}",
        },
    )


@router.get("/invoice/{invoice_id}/pdf")
async def export_invoice_pdf(
    invoice_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(jwt_utils.get_current_user),
):
    """Генерира PDF на фактура.
    """
    import logging
    logger = logging.getLogger(__name__)

    try:
        result = await db.execute(
            select(Invoice)
            .options(selectinload(Invoice.supplier))
            .where(Invoice.id == invoice_id),
        )
        invoice = result.scalars().first()
        if not invoice:
            raise HTTPException(status_code=404, detail="Invoice not found")

        if invoice.company_id != current_user.company_id:
            raise HTTPException(status_code=403, detail="Нямате достъп до тази фактура")

        items_result = await db.execute(
            select(InvoiceItem)
            .options(selectinload(InvoiceItem.ingredient))
            .where(InvoiceItem.invoice_id == invoice_id),
        )
        items = items_result.scalars().all()

        company_result = await db.execute(select(Company).where(Company.id == current_user.company_id))
        company = company_result.scalars().first()

        logger.info(f"Generating PDF for invoice {invoice_id}: number={invoice.number}, items_count={len(items)}, total={invoice.total}, company={company.name if company else None}")

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=40, rightMargin=40, topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()

        # Създай персонални стилове
        styles.add(styles["Normal"].clone("InvoiceTitle"))
        styles["InvoiceTitle"].fontName = BOLD_FONT
        styles["InvoiceTitle"].fontSize = 16
        styles["InvoiceTitle"].spaceAfter = 6

        styles.add(styles["Normal"].clone("InvoiceText"))
        styles["InvoiceText"].fontName = DEFAULT_FONT
        styles["InvoiceText"].fontSize = 9
        styles["InvoiceText"].spaceAfter = 1

        styles.add(styles["Normal"].clone("InvoiceSmall"))
        styles["InvoiceSmall"].fontName = DEFAULT_FONT
        styles["InvoiceSmall"].fontSize = 8
        styles["InvoiceSmall"].spaceAfter = 1

        for style_name in styles.byName:
            style = styles[style_name]
            if "Bold" in style_name or "Title" in style_name or "Heading" in style_name:
                style.fontName = BOLD_FONT
            else:
                style.fontName = DEFAULT_FONT

        # Заглавна част - ФАКТУРА
        elements.append(Paragraph("ФАКТУРА", styles["InvoiceTitle"]))

        # Номер и дата в една линия
        invoice_info = f"№ {invoice.number or invoice.id}                              Дата: {invoice.date.strftime('%d.%m.%Y') if invoice.date else 'N/A'}"
        elements.append(Paragraph(invoice_info, styles["InvoiceText"]))
        elements.append(Paragraph(f"<b>{invoice.griff or 'ОРИГИНАЛ'}</b>", styles["InvoiceText"]))

        elements.append(Spacer(1, 10))

        # ИЗДАТЕЛ и ПОЛУЧАТЕЛ
        if invoice.type == "incoming":
            issuer_name = invoice.supplier.name if invoice.supplier else "-"
            issuer_address = getattr(invoice.supplier, "address", "-") if invoice.supplier else "-"
            issuer_eik = getattr(invoice.supplier, "eik", "-") if invoice.supplier else "-"
            issuer_vat = getattr(invoice.supplier, "vat_number", "-") if invoice.supplier else "-"

            recipient_name = company.name if company else "-"
            recipient_address = company.address if company else "-"
            recipient_eik = company.eik if company else "-"
            recipient_vat = getattr(company, "vat_number", "-") if company else "-"
        else:
            issuer_name = company.name if company else "-"
            issuer_address = company.address if company else "-"
            issuer_eik = company.eik if company else "-"
            issuer_vat = getattr(company, "vat_number", "-") if company else "-"

            recipient_name = invoice.client_name or "-"
            recipient_address = invoice.client_address or "-"
            recipient_eik = invoice.client_eik or "-"
            recipient_vat = getattr(invoice, "client_vat_number", "-") or "-"

        # Таблица за ИЗДАТЕЛ и ПОЛУЧАТЕЛ - СЪС СЪЩАТА ШИРИНА КАТО ТАБЛИЦАТА С АРТИКУЛИ
        table_width = 490  # Обща ширина на страницата

        party_table_data = [
            [Paragraph("<b>ИЗДАТЕЛ</b>", styles["InvoiceSmall"]), Paragraph("<b>ПОЛУЧАТЕЛ</b>", styles["InvoiceSmall"])],
            [Paragraph(f"<b>{issuer_name}</b>", styles["InvoiceSmall"]), Paragraph(f"<b>{recipient_name}</b>", styles["InvoiceSmall"])],
            [Paragraph(f"Адрес: {issuer_address}", styles["InvoiceSmall"]), Paragraph(f"Адрес: {recipient_address}", styles["InvoiceSmall"])],
            [Paragraph(f"ЕИК: {issuer_eik}", styles["InvoiceSmall"]), Paragraph(f"ЕИК: {recipient_eik}", styles["InvoiceSmall"])],
            [Paragraph(f"ИН по ЗДДС: {issuer_vat}", styles["InvoiceSmall"]), Paragraph(f"ИН по ЗДДС: {recipient_vat}", styles["InvoiceSmall"])],
        ]

        half_width = table_width // 2
        party_table = Table(party_table_data, colWidths=[half_width, table_width - half_width])
        party_table_bg = colors.Color(0.92, 0.92, 0.92)
        party_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), BOLD_FONT),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, -1), party_table_bg),
            ("BOX", (0, 0), (-1, -1), 1, party_table_bg),
            ("LINEBEFORE", (1, 0), (1, -1), 1, party_table_bg),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elements.append(party_table)

        elements.append(Spacer(1, 10))

        # Таблица с артикули - СЪС СЪЩАТА ШИРИНА
        vat_rate = float(invoice.vat_rate) if invoice.vat_rate else 20

        # Колонни ширини за таблицата с артикули
        col_widths = [18, 120, 42, 38, 55, 55, 52, 50, 30, 30]  # Общо 490

        table_data = [
            [Paragraph("<b>№</b>", styles["InvoiceSmall"]),
             Paragraph("<b>Наименование</b>", styles["InvoiceSmall"]),
             Paragraph("<b>К-во</b>", styles["InvoiceSmall"]),
             Paragraph("<b>Ед.мярка</b>", styles["InvoiceSmall"]),
             Paragraph("<b>Цена без ДДС</b>", styles["InvoiceSmall"]),
             Paragraph("<b>Цена със ДДС</b>", styles["InvoiceSmall"]),
             Paragraph("<b>Партида</b>", styles["InvoiceSmall"]),
             Paragraph("<b>Годност</b>", styles["InvoiceSmall"]),
             Paragraph("<b>ДДС %</b>", styles["InvoiceSmall"]),
             Paragraph("<b>Сума</b>", styles["InvoiceSmall"])],
        ]

        for idx, item in enumerate(items, 1):
            item_name = item.name or (item.ingredient.name if item.ingredient else f"Артикул {item.id}")
            unit_price = float(item.unit_price) if item.unit_price else 0
            unit_price_vat = float(item.unit_price_with_vat) if item.unit_price_with_vat else None
            batch_num = item.batch_number or "-"
            expiry = item.expiration_date.strftime("%d.%m.%Y") if item.expiration_date else "-"
            item_total = float(item.total) if item.total else 0

            table_data.append([
                Paragraph(str(idx), styles["InvoiceSmall"]),
                Paragraph(item_name[:30], styles["InvoiceSmall"]),
                Paragraph(f"{float(item.quantity):.3f}" if item.quantity else "-", styles["InvoiceSmall"]),
                Paragraph(item.unit or "бр.", styles["InvoiceSmall"]),
                Paragraph(f"{unit_price:.2f}" if unit_price else "-", styles["InvoiceSmall"]),
                Paragraph(f"{unit_price_vat:.2f}" if unit_price_vat else "-", styles["InvoiceSmall"]),
                Paragraph(batch_num[:15], styles["InvoiceSmall"]),
                Paragraph(expiry[:12], styles["InvoiceSmall"]),
                Paragraph(f"{vat_rate:.0f}%", styles["InvoiceSmall"]),
                Paragraph(f"{item_total:.2f}" if item_total else "-", styles["InvoiceSmall"]),
            ])

        # Крайни суми
        subtotal = float(invoice.subtotal) if invoice.subtotal else 0
        vat_amount = float(invoice.vat_amount) if invoice.vat_amount else 0
        total_amount = float(invoice.total) if invoice.total else 0

        table_data.append([
            "", "", "", "", "", "", "", "",
            Paragraph("<b>Сума:</b>", styles["InvoiceSmall"]),
            Paragraph(f"<b>{subtotal:.2f}</b>", styles["InvoiceSmall"]),
        ])
        table_data.append([
            "", "", "", "", "", "", "", "",
            Paragraph(f"<b>ДДС {vat_rate:.0f}%:</b>", styles["InvoiceSmall"]),
            Paragraph(f"<b>{vat_amount:.2f}</b>", styles["InvoiceSmall"]),
        ])
        table_data.append([
            "", "", "", "", "", "", "", "",
            Paragraph("<b>ОБЩО:</b>", styles["InvoiceSmall"]),
            Paragraph(f"<b>{total_amount:.2f} €</b>", styles["InvoiceSmall"]),
        ])

        items_table = Table(table_data, colWidths=col_widths)

        # Цвят за header и border
        header_bg = colors.Color(0.85, 0.85, 0.85)

        # Стил за таблицата с артикули
        table_style = [
            ("FONTNAME", (0, 0), (-1, 0), BOLD_FONT),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, 0), header_bg),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("ALIGN", (1, 0), (1, -1), "LEFT"),
            ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -2), 0.5, header_bg),
            ("LINEBELOW", (0, -3), (-1, -3), 1, header_bg),
            ("BOX", (0, 0), (-1, -1), 1, header_bg),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]

        # Сив фон за последните 3 реда
        for i in range(1, 4):
            row_idx = -i
            table_style.append(("BACKGROUND", (8, row_idx), (9, row_idx), colors.Color(0.92, 0.92, 0.92)))

        items_table.setStyle(TableStyle(table_style))
        elements.append(items_table)

        elements.append(Spacer(1, 10))

        # Дати и плащане
        payment_info = []
        if invoice.payment_method:
            payment_info.append(f"Начин на плащане: {invoice.payment_method}")
        if invoice.due_date:
            payment_info.append(f"Срок на плащане: {invoice.due_date.strftime('%d.%m.%Y')}")
        if getattr(invoice, "payment_date", None):
            payment_info.append(f"Дата на плащане: {invoice.payment_date.strftime('%d.%m.%Y')}")

        if payment_info:
            elements.append(Paragraph(" | ".join(payment_info), styles["InvoiceText"]))

        if invoice.notes:
            elements.append(Spacer(1, 6))
            elements.append(Paragraph(f"<b>Бележки:</b> {invoice.notes}", styles["InvoiceSmall"]))

        elements.append(Spacer(1, 15))

        # Подписи с QR код ПОД тях
        signatures_data = [
            [Paragraph("_" * 45, styles["InvoiceSmall"]), Paragraph("_" * 35, styles["InvoiceSmall"])],
            [Paragraph("За издателя", styles["InvoiceSmall"]), Paragraph("За получателя", styles["InvoiceSmall"])],
        ]
        signatures_table = Table(signatures_data, colWidths=[245, 245])
        signatures_table.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))

        elements.append(signatures_table)

        elements.append(Spacer(1, 8))

        # QR код - ПОД подписите, по-малък
        try:
            qr = qrcode.QRCode(version=1, box_size=8, border=3)
            qr.add_data(f"BULSTAT:{issuer_eik};INV:{invoice.number};DATE:{invoice.date.strftime('%Y%m%d') if invoice.date else ''};SUM:{total_amount:.2f}")
            qr.make(fit=True)
            qr_img = qr.make_image(fill_color="black", back_color="white")

            qr_buffer = BytesIO()
            qr_img.save(qr_buffer, format="PNG")
            qr_buffer.seek(0)

            # QR код с 10% по-малък (63 вместо 70)
            qr_image = Image(qr_buffer, width=63, height=63)

            # QR кода центриран
            qr_table_data = [
                [qr_image],
                [Paragraph(f"QR код: {invoice.number}", styles["InvoiceSmall"])],
            ]
            qr_table = Table(qr_table_data, colWidths=[63])
            qr_table.setStyle(TableStyle([
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ]))
            elements.append(qr_table)
        except Exception as qr_error:
            logger.warning(f"Failed to generate QR code: {qr_error}")

        doc.build(elements)

        buffer.seek(0)

        # Only use ASCII for filename header
        ascii_filename = f"invoice_{invoice_id}.pdf"
        encoded_filename = quote(ascii_filename)

        return Response(
            content=buffer.getvalue(),
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=\"{ascii_filename}\"; filename*=UTF-8''{encoded_filename}",
            },
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating PDF for invoice {invoice_id}: {e!s}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Грешка при генериране на PDF: {e!s}") from e


@router.get("/cash-journal/xlsx")
async def export_cash_journal_xlsx(
    start_date: str | None = None,
    end_date: str | None = None,
    operation_type: str | None = None,
    payment_method: str | None = None,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(jwt_utils.get_current_user),
):
    """Export cash journal entries to XLSX"""
    from backend.database.models import CashJournalEntry

    stmt = select(CashJournalEntry)
    if current_user.role.name != "super_admin":
        stmt = stmt.where(CashJournalEntry.company_id == current_user.company_id)
    if start_date:
        stmt = stmt.where(CashJournalEntry.date >= datetime.date.fromisoformat(start_date))
    if end_date:
        stmt = stmt.where(CashJournalEntry.date <= datetime.date.fromisoformat(end_date))
    if operation_type:
        stmt = stmt.where(CashJournalEntry.operation_type == operation_type)
    if payment_method:
        stmt = stmt.where(CashJournalEntry.payment_method == payment_method)

    stmt = stmt.order_by(CashJournalEntry.date.asc(), CashJournalEntry.id.asc())
    result = await db.execute(stmt)
    entries = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Касов дневник"

    # Header style
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    income_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    expense_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    date_format = "DD.MM.YYYY"
    number_format = '#,##0.00'

    # Headers
    headers = ["Дата", "Тип", "Сума", "Платежен метод", "Описание", "Източник", "Фактура №", "Създал"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    payment_method_labels = {
        "cash": "В брой",
        "bank_transfer": "Банков превод",
        "card": "Карта",
        "other": "Друго",
    }

    total_income = 0
    total_expense = 0

    for i, entry in enumerate(entries, 2):
        ws.cell(row=i, column=1, value=entry.date).number_format = date_format

        op_type = "Приход" if entry.operation_type == "income" else "Разход"
        ws.cell(row=i, column=2, value=op_type)

        ws.cell(row=i, column=3, value=float(entry.amount)).number_format = number_format
        if entry.operation_type == "income":
            ws.cell(row=i, column=3).fill = income_fill
            total_income += float(entry.amount)
        else:
            ws.cell(row=i, column=3).fill = expense_fill
            total_expense += float(entry.amount)

        ws.cell(row=i, column=4, value=payment_method_labels.get(entry.payment_method, entry.payment_method or "-"))
        ws.cell(row=i, column=5, value=entry.description or "-")

        source = "Ръчна" if entry.reference_type != "invoice" else "Фактура"
        ws.cell(row=i, column=6, value=source)

        if entry.reference_type == "invoice" and entry.reference_id:
            inv_stmt = select(Invoice).where(Invoice.id == entry.reference_id)
            inv_result = await db.execute(inv_stmt)
            inv = inv_result.scalars().first()
            ws.cell(row=i, column=7, value=inv.number if inv else "-")
        else:
            ws.cell(row=i, column=7, value="-")

        if entry.created_by:
            user_stmt = select(User).where(User.id == entry.created_by)
            user_result = await db.execute(user_stmt)
            user = user_result.scalars().first()
            ws.cell(row=i, column=8, value=f"{user.first_name or ''} {user.last_name or ''}".strip() if user else "-")
        else:
            ws.cell(row=i, column=8, value="-")

    # Summary row
    summary_row = len(entries) + 3
    ws.cell(row=summary_row, column=1, value="ОБОБЩЕНИЕ").font = Font(bold=True, size=12)
    ws.cell(row=summary_row + 1, column=2, value="Общо приходи:").font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=3, value=total_income).number_format = number_format
    ws.cell(row=summary_row + 1, column=3).font = Font(bold=True, color="2E7D32")
    ws.cell(row=summary_row + 2, column=2, value="Общо разходи:").font = Font(bold=True)
    ws.cell(row=summary_row + 2, column=3, value=total_expense).number_format = number_format
    ws.cell(row=summary_row + 2, column=3).font = Font(bold=True, color="C62828")
    ws.cell(row=summary_row + 3, column=2, value="Салдо:").font = Font(bold=True, size=12)
    ws.cell(row=summary_row + 3, column=3, value=total_income - total_expense).number_format = number_format
    ws.cell(row=summary_row + 3, column=3).font = Font(bold=True, size=12)

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 20

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"kasov-dnevnik-{start_date or 'all'}-to-{end_date or 'all'}.xlsx"
    encoded_filename = quote(filename)

    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=\"{filename}\"; filename*=UTF-8''{encoded_filename}",
        },
    )


@router.get("/payment-batch/{batch_id}/csv")
async def export_payment_batch_csv(
    batch_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(jwt_utils.get_current_user),
):
    if current_user.role.name not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Само администратори могат да експортират платежни батчове")

    from backend.services.salary_payment_service import SalaryPaymentService

    service = SalaryPaymentService(db)
    csv_content = await service.export_batch(batch_id, format="csv")

    filename = f"payment-batch-{batch_id}.csv"
    encoded_filename = quote(filename)

    return Response(
        content=csv_content,
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f"attachment; filename=\"{filename}\"; filename*=UTF-8''{encoded_filename}",
        },
    )

