"""Fleet module export endpoints."""
import datetime
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from backend.auth.module_guard import require_module_dep
from backend.database.database import get_db
from backend.database.models import (
    Vehicle,
    VehicleFuel,
    VehicleRepair,
    VehicleInsurance,
    VehicleInspection,
    VehicleMileage,
    VehicleVignette,
    VehicleToll,
    VehicleDriver,
    VehicleType,
)
from backend.auth import jwt_utils
from sqlalchemy.orm import selectinload

router = APIRouter(
    prefix="/api/fleet",
    tags=["fleet-export"],
    dependencies=[Depends(require_module_dep("fleet"))],
)

# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
CURRENCY_FORMAT = '#,##0.00 "лв"'
DATE_FORMAT = 'DD.MM.YYYY'


def style_header_row(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def style_data_cell(cell, is_currency=False, is_date=False):
    cell.border = THIN_BORDER
    if is_currency:
        cell.number_format = CURRENCY_FORMAT
    if is_date:
        cell.number_format = DATE_FORMAT


@router.get("/export/vehicles")
async def export_fleet_vehicles(
    db: AsyncSession = Depends(get_db),
    current_user = Depends(jwt_utils.get_current_user),
):
    """Export fleet vehicles and their costs to Excel."""
    
    # Fetch all vehicles with relationships
    stmt = select(Vehicle).options(
        selectinload(Vehicle.vehicle_type),
        selectinload(Vehicle.fuel_cards),
    ).order_by(Vehicle.registration_number)
    
    result = await db.execute(stmt)
    vehicles = result.scalars().all()

    wb = Workbook()
    
    # Sheet 1: Vehicles Overview
    ws_vehicles = wb.active
    ws_vehicles.title = "Автомобили"
    
    headers_vehicles = [
        "Рег. номер", "VIN", "Марка", "Модел", "Година", "Тип", "Гориво", 
        "Статус", "Начален пробег", "Собственик", "Забележки"
    ]
    
    for col, header in enumerate(headers_vehicles, 1):
        ws_vehicles.cell(row=1, column=col, value=header)
    style_header_row(ws_vehicles, len(headers_vehicles))
    
    for row_idx, v in enumerate(vehicles, 2):
        ws_vehicles.cell(row=row_idx, column=1, value=v.registration_number)
        ws_vehicles.cell(row=row_idx, column=2, value=v.vin)
        ws_vehicles.cell(row=row_idx, column=3, value=v.make)
        ws_vehicles.cell(row=row_idx, column=4, value=v.model)
        ws_vehicles.cell(row=row_idx, column=5, value=v.year)
        ws_vehicles.cell(row=row_idx, column=6, value=v.vehicle_type.name if v.vehicle_type else "-")
        ws_vehicles.cell(row=row_idx, column=7, value=v.fuel_type)
        ws_vehicles.cell(row=row_idx, column=8, value=v.status)
        ws_vehicles.cell(row=row_idx, column=9, value=v.initial_mileage)
        ws_vehicles.cell(row=row_idx, column=10, value=v.owner_name or "Фирмен")
        ws_vehicles.cell(row=row_idx, column=11, value=v.notes)
        
        for col in range(1, len(headers_vehicles) + 1):
            style_data_cell(ws_vehicles.cell(row=row_idx, column=col))

    # Sheet 2: Cost Summary
    ws_costs = wb.create_sheet("Разходи")
    
    headers_costs = [
        "Рег. номер", "Марка/Модел", "Гориво (лв)", "Ремонти (лв)", 
        "ГТП (лв)", "Застраховки (лв)", "Винетки (лв)", "Такси (лв)", 
        "ОБЩО (лв)", "Пробег (км)", "Лв/км"
    ]
    
    for col, header in enumerate(headers_costs, 1):
        ws_costs.cell(row=1, column=col, value=header)
    style_header_row(ws_costs, len(headers_costs))
    
    for row_idx, v in enumerate(vehicles, 2):
        # Calculate costs
        fuel_stmt = select(VehicleFuel.total).where(VehicleFuel.vehicle_id == v.id)
        fuel_res = await db.execute(fuel_stmt)
        total_fuel = sum([f for f in fuel_res.scalars().all() if f]) or 0
        
        repair_stmt = select(VehicleRepair.cost).where(VehicleRepair.vehicle_id == v.id)
        repair_res = await db.execute(repair_stmt)
        total_repairs = sum([r for r in repair_res.scalars().all() if r]) or 0
        
        insp_stmt = select(VehicleInspection.cost).where(VehicleInspection.vehicle_id == v.id)
        insp_res = await db.execute(insp_stmt)
        total_insp = sum([i for i in insp_res.scalars().all() if i]) or 0
        
        ins_stmt = select(VehicleInsurance.premium).where(VehicleInsurance.vehicle_id == v.id)
        ins_res = await db.execute(ins_stmt)
        total_ins = sum([i for i in ins_res.scalars().all() if i]) or 0
        
        vig_stmt = select(VehicleVignette.cost).where(VehicleVignette.vehicle_id == v.id)
        vig_res = await db.execute(vig_stmt)
        total_vig = sum([v for v in vig_res.scalars().all() if v]) or 0
        
        toll_stmt = select(VehicleToll.cost).where(VehicleToll.vehicle_id == v.id)
        toll_res = await db.execute(toll_stmt)
        total_toll = sum([t for t in toll_res.scalars().all() if t]) or 0
        
        grand_total = total_fuel + total_repairs + total_insp + total_ins + total_vig + total_toll
        
        # Current mileage
        mile_stmt = select(VehicleMileage.mileage).where(VehicleMileage.vehicle_id == v.id).order_by(VehicleMileage.date.desc()).limit(1)
        mile_res = await db.execute(mile_stmt)
        current_mileage = mile_res.scalars().first() or v.initial_mileage
        distance = current_mileage - v.initial_mileage
        
        cost_per_km = (grand_total / distance) if distance > 0 else 0
        
        ws_costs.cell(row=row_idx, column=1, value=v.registration_number)
        ws_costs.cell(row=row_idx, column=2, value=f"{v.make} {v.model}")
        ws_costs.cell(row=row_idx, column=3, value=total_fuel)
        ws_costs.cell(row=row_idx, column=4, value=total_repairs)
        ws_costs.cell(row=row_idx, column=5, value=total_insp)
        ws_costs.cell(row=row_idx, column=6, value=total_ins)
        ws_costs.cell(row=row_idx, column=7, value=total_vig)
        ws_costs.cell(row=row_idx, column=8, value=total_toll)
        ws_costs.cell(row=row_idx, column=9, value=grand_total)
        ws_costs.cell(row=row_idx, column=10, value=distance)
        ws_costs.cell(row=row_idx, column=11, value=cost_per_km)
        
        for col in range(1, len(headers_costs) + 1):
            is_curr = col in [3, 4, 5, 6, 7, 8, 9]
            style_data_cell(ws_costs.cell(row=row_idx, column=col), is_currency=is_curr)

    # Adjust column widths
    for ws in [ws_vehicles, ws_costs]:
        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 40)

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"Fleet_Export_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
